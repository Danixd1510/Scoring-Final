import pdfplumber
import re
from datetime import datetime
from mapping import MAPEO_CASILLAS

def limpiar_valor(valor):
    if valor is None: return 0
    # Limpia: espacios, comas, paréntesis a guion
    # Ejemplo: "(1,234.00)" -> "-1234.00"
    clean = str(valor).strip().replace(' ', '').replace(',', '')
    if '(' in clean or ')' in clean:
        clean = '-' + clean.replace('(', '').replace(')', '')
    try:
        return float(clean)
    except:
        return 0

def extraer_datos_de_pdf(archivo_pdf):
    datos_extraidos = {}
    with pdfplumber.open(archivo_pdf) as pdf:
        for page in pdf.pages:
            # layout=True es vital: mantiene la estructura visual
            text = page.extract_text(layout=True)
            if not text: continue
            
            lineas = text.split('\n')
            for linea in lineas:
                # Buscamos cada código en la línea
                for code in MAPEO_CASILLAS.keys():
                    # Buscamos el código seguido de cualquier espacio y luego un número
                    # El regex busca: código, espacios, y cualquier secuencia de dígitos, puntos, menos o paréntesis
                    patron = rf"\b{code}\b\s+([-\d\.\,\(\)\s]+)"
                    match = re.search(patron, linea)
                    
                    if match:
                        valor_encontrado = match.group(1).strip()
                        # Solo asignamos si es un número real (que contenga al menos un dígito)
                        if any(c.isdigit() for c in valor_encontrado):
                            datos_extraidos[code] = limpiar_valor(valor_encontrado)
                            
    return datos_extraidos

import streamlit as st
import openpyxl
from utils import extraer_ficha_ruc

st.set_page_config(page_title="Extractor RUC", layout="centered")
st.title("Extractor de Ficha RUC")

ficha_file = st.file_uploader("Subir Ficha RUC (PDF)", type=["pdf"])

if ficha_file:
    # Extraer info
    info = extraer_ficha_ruc(ficha_file)
    
    # Mostrar en pantalla
    st.success("Información extraída correctamente:")
    st.write(f"**Empresa:** {info['Nombre']}")
    st.write(f"**RUC:** {info['RUC']}")
    st.write(f"**Inicio de Actividades:** {info['Inicio']}")
    
    # Botón para actualizar el Excel
    if st.button("Actualizar Scoring Final.xlsx"):
        wb = openpyxl.load_workbook("Scoring Final.xlsx")
        
        # Escribir en la pestaña SCORING_FINAL
        if "SCORING_FINAL" in wb.sheetnames:
            ws = wb["SCORING_FINAL"]
            ws["C2"] = info['Nombre'].replace(".", "") # Cliente
            ws["C3"] = info['RUC']                    # RUC
            ws["C4"] = info['Inicio']                 # Fecha inicio
            
            nombre_final = f"Scoring Final - {info['Nombre'].replace('.', '')}.xlsx"
            wb.save(nombre_final)
            
            with open(nombre_final, "rb") as f:
                st.download_button("📥 Descargar Excel", f, file_name=nombre_final)
