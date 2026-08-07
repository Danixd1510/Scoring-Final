import streamlit as st
import openpyxl
import os
import io
import re
from pathlib import Path
from utils import extraer_datos_de_pdf, extraer_ficha_ruc, extraer_reporte_tributario
from mapping import MAPEO_CASILLAS

st.set_page_config(page_title="Scoring Financiero", layout="wide")
st.title("SACA TUS RATIOS RAPIDISIMOOOO")

def sanitize_filename(name: str) -> str:
    if not name:
        return "cliente"
    clean = re.sub(r'[\/\\\:\*\?"<>\|]', "_", name)
    clean = re.sub(r"[^\w\-\s]", "_", clean)
    clean = clean.strip()
    if not clean:
        return "cliente"
    return clean

def to_filelike(uploaded_file):
    if uploaded_file is None:
        return None
    if hasattr(uploaded_file, "read"):
        data = uploaded_file.read()
        return io.BytesIO(data)
    return uploaded_file

# --- VISTA EN PANTALLA ---
col1, col2 = st.columns(2)
ficha_file = col1.file_uploader("Subir Ficha RUC", type=["pdf"])
reporte_file = col2.file_uploader("Subir Reporte Tributario", type=["pdf"])

if ficha_file:
    ficha_stream = to_filelike(ficha_file)
    try:
        info = extraer_ficha_ruc(ficha_stream)
        st.success(f"**Empresa:** {info.get('Nombre','-')} | **RUC:** {info.get('RUC','-')} | **Inicio:** {info.get('Inicio','-')}")
    except Exception as e:
        st.error(f"No se pudo extraer ficha RUC: {e}")

if reporte_file:
    reporte_stream = to_filelike(reporte_file)
    try:
        ventas, mes = extraer_reporte_tributario(reporte_stream)
        st.info(f"**Ventas Totales:** S/ {ventas:,.2f} | **Mes detectado:** {mes if mes else 'N/D'}")
    except Exception as e:
        st.error(f"No se pudo extraer reporte tributario: {e}")

st.divider()

# --- GENERAR EXCEL (usa siempre la plantilla Scoring Final.xlsx del repo) ---
st.header("Generar Reporte de Ratios")
cliente = st.text_input("Nombre del Cliente")

st.write("La aplicación usa la plantilla 'Scoring Final.xlsx' incluida en el repositorio.")

pdf_2023 = st.file_uploader("Subir PDF 2023", type=["pdf"])
pdf_2024 = st.file_uploader("Subir PDF 2024", type=["pdf"])
pdf_2025 = st.file_uploader("Subir PDF 2025", type=["pdf"])

archivos = {"2023": pdf_2023, "2024": pdf_2024, "2025": pdf_2025}

if st.button("Generar Excel"):
    if not cliente:
        st.error("Error: Debes ingresar el nombre del cliente.")
    else:
        # Buscar la plantilla junto al archivo app.py (más robusto en despliegues)
        base_path = Path(__file__).parent
        plantilla_path = base_path / "Scoring Final.xlsx"

        if not plantilla_path.exists():
            st.error(f"Error: No se encontró la plantilla '{plantilla_path.name}' en el repositorio.")
            st.stop()

        try:
            with open(plantilla_path, "rb") as f:
                wb = openpyxl.load_workbook(f)
        except Exception as e:
            st.error(f"No se pudo cargar la plantilla: {e}")
            st.stop()

        total_reemplazadas = 0
        detalles_por_anio = {}
        claves_sin_mapeo = set()

        for anio, archivo in archivos.items():
            reemplazadas = 0
            if archivo is None:
                detalles_por_anio[anio] = {"reemplazadas": 0, "mensaje": "No se subió archivo"}
                continue

            if anio not in wb.sheetnames:
                detalles_por_anio[anio] = {"reemplazadas": 0, "mensaje": f"Hoja '{anio}' no encontrada en plantilla"}
                continue

            ws = wb[anio]
            stream = to_filelike(archivo)
            try:
                datos = extraer_datos_de_pdf(stream)
            except Exception as e:
                detalles_por_anio[anio] = {"reemplazadas": 0, "mensaje": f"Error extrayendo PDF: {e}"}
                continue

            for casilla, valor in datos.items():
                celda = MAPEO_CASILLAS.get(casilla)
                if not celda:
                    claves_sin_mapeo.add(casilla)
                    continue

                cell_obj = ws[celda]
                has_formula = False
                try:
                    if getattr(cell_obj, "data_type", None) == "f":
                        has_formula = True
                except:
                    pass
                cur_val = cell_obj.value
                if isinstance(cur_val, str) and cur_val.startswith("="):
                    has_formula = True

                if has_formula:
                    continue

                try:
                    if isinstance(valor, (int, float)):
                        cell_obj.value = float(valor)
                    else:
                        cell_obj.value = float(valor)
                except:
                    cell_obj.value = valor

                try:
                    cell_obj.number_format = '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"??_);_(@_)'
                except:
                    pass

                reemplazadas += 1

            detalles_por_anio[anio] = {"reemplazadas": reemplazadas, "mensaje": "OK" if reemplazadas else "No se reemplazaron celdas"}
            total_reemplazadas += reemplazadas

        nombre_final = f"Scoring Final - {sanitize_filename(cliente)}.xlsx"
        buffer = io.BytesIO()
        try:
            wb.save(buffer)
            buffer.seek(0)
            st.success(f"Generado: {nombre_final} — celdas modificadas: {total_reemplazadas}")
            for anio, info in detalles_por_anio.items():
                st.write(f"{anio}: {info['reemplazadas']} celdas - {info['mensaje']}")
            if claves_sin_mapeo:
                st.warning(f"Claves encontradas en PDFs sin mapeo en MAPEO_CASILLAS: {', '.join(sorted(claves_sin_mapeo))}")
            st.download_button("📥 Descargar Excel", data=buffer, file_name=nombre_final, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"No se pudo guardar el archivo final: {e}")
